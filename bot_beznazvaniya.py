from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import asyncio
from imaplib import Commands
import logging
import os
import pandas as pd
import calendar
from aiogram import Bot, types
from aiogram.utils.markdown import text, bold, italic, code, pre, hitalic, escape_md, _join, hbold, hcode, hpre, underline, hunderline, strikethrough, hstrikethrough, link, hlink, hide_link
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import InputFile
from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton
from time import sleep
from aiogram.types import ParseMode, InputMediaPhoto, InputMediaVideo, ChatActions
from aiogram.utils import executor
from aiogram.dispatcher import Dispatcher
from aiogram.types.message import ContentType
import logging
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters import Text
import aiogram.utils.markdown as fmt
from openpyxl.styles import colors
from openpyxl.styles import Font, Color


bot = Bot(token='5774818988:AAGwwpEM0EgTYBQWdrmgdlr3dEZgXxvSmQE')
dp = Dispatcher(bot, storage=MemoryStorage())
admin_chat_id = -967480953

logging.basicConfig(level=logging.INFO)


start_button = types.ReplyKeyboardMarkup(resize_keyboard=True)
info = types.KeyboardButton("Регистрация")
stats = types.KeyboardButton("Оплата")
call = types.KeyboardButton("Наши контакты")
start_button.add(info, stats)
start_button.add(call)

V_button = types.ReplyKeyboardMarkup(resize_keyboard=True)
V1 = types.KeyboardButton("1299 руб. с репостом афиши из группы VK к себе на страницу")
V2 = types.KeyboardButton("1499 руб. без репоста")
cancle = types.KeyboardButton('Отмена')
V_button.add(V1)
V_button.add(V2)
V_button.add(cancle)



cancle_button = types.ReplyKeyboardMarkup(resize_keyboard=True)
cancle = types.KeyboardButton('Отмена')
cancle_button.add(cancle)


@dp.message_handler(commands='start')
async def start(message: types.Message):
    await message.answer(f'Салют! \n\n' + bold('BEZNAZVANIYA') + ' | ' + bold('ACT A FOOL DAY') + ' 🤹🏽‍♂️\n' + '31 МАРТА - 1 АПРЕЛЯ\n23:00 - 5:00\nЛОФТ ФАРФОР\n\n'+ bold('FREE BAR') + ' | ' + bold('PRANKS') + ' | ' + bold("INVATED DJS")+ ' | ' + bold('FC') + '/' + bold('DC') + bold('18') + '+' + '\n\nМеню:\n/enroll - пройти регистрацию\n/payment - оплатить билет\n/contacts - наши контакты', parse_mode=ParseMode.MARKDOWN, reply_markup=start_button)# \n/question - связаться с нами


@dp.message_handler(Text(equals="Отмена"), state="*")
async def menu_button(message: types.Message, state: FSMContext):
    await state.finish()
    await bot.send_message(
        message.chat.id, "Отмена произошла успешно.")
    await message.answer(f'Салют! \n\n' + bold('BEZNAZVANIYA') + ' | ' + bold('ACT A FOOL DAY') + ' 🤹🏽‍♂️\n' + '31 МАРТА - 1 АПРЕЛЯ\n23:00 - 5:00\nЛОФТ ФАРФОР\n\n'+ bold('FREE BAR') + ' | ' + bold('PRANKS') + ' | ' + bold("INVATED DJS") +' | ' + bold('FC') + '/' + bold('DC') + bold('18') + '+' + '\n\nМеню:\n/enroll - пройти регистрацию\n/payment - оплатить билет\n/contacts - наши контакты', parse_mode=ParseMode.MARKDOWN, reply_markup=start_button)# \n/question - связаться с нами
    await state.finish()


class meinfo(StatesGroup):
    Q1 = State()
    Q2 = State()
    Q3 = State()
    Q4 = State()


class PGroup(StatesGroup):
    W1 = State()
    W2 = State()
    W3 = State()
    W4 = State()
    W5 = State()


@dp.message_handler(commands='payment', state=None)
async def enter_meinfo(message: types.Message):
    await message.answer(f'Выберите вариант оплаты', parse_mode=ParseMode.MARKDOWN, reply_markup=V_button)# \n/question - связаться с нами

    # J = open("/Users/senyashago/Desktop/Бот практика/bez/check.txt","r", encoding="utf-8") 
    # joinedUsers = set()
    # for line in J:
    #     joinedUsers.add(line.strip())
    # if str(message.chat.id) in joinedUsers:
    #     text = bold("Введите ФИО, указанное при регистрации") + "\nПример: " + italic("Иванов Иван Иванович")
    #     await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)                                
    #     await PGroup.W1.set()
    # else:
    #     text = "Сначала Вам надо зарегистрироваться. Для этого нажмите команду /enroll или кнопку " + bold('"Регистрация"') + " внизу👇"
    #     await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)


@dp.message_handler(commands='contacts')
async def enter_meinfo(message: types.Message):
    keyboard = InlineKeyboardMarkup()
    url_button1 = InlineKeyboardButton(text="Менеджер VK", url="https://vk.com/beznazvaniya_manager")
    url_button5 = InlineKeyboardButton(text="Наш инстаграм (тут будет розыгрыш мерча)", url="https://instagram.com/beznazvaniya_msk?igshid=YmMyMTA2M2Y=")
    url_button6 = InlineKeyboardButton(text="Группа VK", url="https://vk.com/beznazvaniya_msk")
    url_button7 = InlineKeyboardButton(text="Telegram-канал", url="https://t.me/joinchat/AAAAAEuYaZ8EnD84P-MUXQ")
    keyboard.add(url_button1)
    keyboard.add(url_button5)
    keyboard.add(url_button6)
    keyboard.add(url_button7)
    await bot.send_message(message.chat.id, bold('Контакты'),reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN)  # message.from_user.id - конкретному пользователю




@dp.message_handler(state=PGroup.W1)
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    ww = wb['Количество']  
    k = 0
    check = 0
    for i in range(2,10000):
            if str(ws['C' + str(i)].value) == str(answer):
                check = 1
                if str(ws['G' + str(i)].value) == '1':
                    text = "Вы в списках! Оплата успешно прошла!"
                    await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
                    k = 1
                    await state.finish()
                    break
    if check == 0:
        await message.answer("Этот человек отсутствует в списках. Проверьте ФИО и регистрацию этого человека. Затем повторите попытку", reply_markup=start_button)
        await state.finish()
    if k == 0 and check == 1:
        await state.update_data(answer11=answer)        
        # text = "Оплатите билет:\n\nСтоимость: " + bold("1299 рублей") + "\n\n2202200112477334\nДмитрий Альбертович И.\n\n" + bold('ВАЖНО: ') + 'В комментарии к платежу укажите фамилию\n\n' + "После оплаты отправьте сюда скрин оплаты (не документ, а именно фото)"
        text = "Оплатите билет:\n\nСтоимость: " + bold("1499 рублей") + "\n\n" + bold("СБЕРБАНК") + "\n5469 3300 1323 0903\nЕлизавета Александровна\n\n" + bold('ВАЖНО: ') + 'В комментарии к платежу укажи фамилию\n\n' + "После оплаты пришли в ответ на это сообщение скрин оплаты (не документ, а именно фото)"

        await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
        await PGroup.W2.set()


@dp.message_handler(state=PGroup.W3)
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    ww = wb['Количество']  
    k = 0
    check = 0
    for i in range(2,10000):
            if str(ws['C' + str(i)].value) == str(answer):
                check = 1
                if str(ws['G' + str(i)].value) == '1':
                    text = "Вы в списках! Оплата успешно прошла!"
                    await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
                    k = 1
                    await state.finish()
                    break
    if check == 0:
        await message.answer("Этот человек отсутствует в списках. Проверьте ФИО и регистрацию этого человека. Затем повторите попытку", reply_markup=start_button)
        await state.finish()
    if k == 0 and check == 1:
        await state.update_data(answer11=answer)
        
        text = bold("Отправьте ссылку на свой репост из VK\n") + '(ссылка на группу vk расположена в контактах)'
        await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
        await PGroup.W4.set()      
        # text = "Оплатите билет:\n\nСтоимость: " + bold("1199 рублей") + "\n\n2202200112477334\nДмитрий Альбертович И.\n\n" + bold('ВАЖНО: ') + 'В комментарии к платежу укажите фамилию\n\n' + "После оплаты отправьте сюда скрин оплаты (не документ, а именно фото)"
        # await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
        # await PGroup.W2.set()

@dp.message_handler(state=PGroup.W4)
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    await state.update_data(answer22=answer)
    data = await state.get_data()               
    answer11 = data.get("answer11")
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    ww = wb['Количество']  
    n =str(int(ww['A2'].value) + 2)
   
    
    idf = ''
    k = 0
    check = 0
    for i in range(2,10000):
            if str(ws['C' + str(i)].value) == str(answer11):
                check = 1
                idf = str(i)
                if str(ws['G' + str(i)].value) == '1':
                    text = "Вы в списках! Оплата успешно прошла!"
                    await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
                    k = 1
                    await state.finish()
                    break
    if check == 0:
        await message.answer("Этот человек отсутствует в списках. Проверьте ФИО и регистрацию этого человека. Затем повторите попытку", reply_markup=start_button)
        await state.finish()
    if k == 0 and check == 1:
        ws['H' + idf] = answer
        wb.save(f)
        wb.close()     
        text = "Оплатите билет:\n\nСтоимость: " + bold("1299 рублей") + "\n\n" + bold("СБЕРБАНК") + "\n5469 3300 1323 0903\nЕлизавета Александровна\n\n" + bold('ВАЖНО: ') + 'В комментарии к платежу укажи фамилию\n\n' + "После оплаты пришли в ответ на это сообщение скрин оплаты (не документ, а именно фото)"
        await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
    wb.close() 
    await PGroup.W5.set() 


@dp.message_handler(content_types=["photo"], state=PGroup.W5)
async def answer_q1(message: types.Message, state: FSMContext):
    data = await state.get_data()               
    answer11 = data.get("answer11")
    answer22 = data.get("answer22")
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    z = 0
    for i in range(2,1000):
        if str(ws['C' + str(i)].value) == str(answer11):
            photo = message.photo.pop()
            await photo.download(f'./images/{answer11}.jpg')
            await bot.send_message(admin_chat_id,f"Поступила оплата от {answer11}\nЦена: 1299 руб.\nusername: @{message.from_user.username}\nchatID: {message.chat.id}\nРепост: {answer22}")
            await bot.send_photo(admin_chat_id, photo=open(f'./images/{answer11}.jpg', 'rb'))
            await message.answer("Информация принята. Ожидайте подтверждения в этом боте.", reply_markup=start_button)
            z = 1
            break
    if z == 0:
        await message.answer("Этот человек отсутствует в списках. Проверьте ФИО и повторите попытку", reply_markup=start_button)
    await state.finish()

@dp.message_handler(content_types=["photo"], state=PGroup.W2)
async def answer_q1(message: types.Message, state: FSMContext):
    data = await state.get_data()               
    answer11 = data.get("answer11")
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    z = 0
    for i in range(2,1000):
        if str(ws['C' + str(i)].value) == str(answer11):
            photo = message.photo.pop()
            await photo.download(f'./images/{answer11}.jpg')
            await bot.send_message(admin_chat_id,f"Поступила оплата от {answer11}\nЦена: 1499 руб.\nusername: @{message.from_user.username}\nchatID: {message.chat.id}")
            await bot.send_photo(admin_chat_id, photo=open(f'./images/{answer11}.jpg', 'rb'))
            await message.answer("Информация принята. Ожидайте подтверждения в этом боте.", reply_markup=start_button)
            z = 1
            break
    if z == 0:
        await message.answer("Этот человек отсутствует в списках. Проверьте ФИО и повторите попытку", reply_markup=start_button)
    await state.finish()
    

 
@dp.message_handler(commands='enroll', state=None)        
async def enter_meinfo(message: types.Message):
    text = bold("Введите ФИО") + "\nПример: " + italic("Иванов Иван Иванович")
    await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
    await meinfo.Q1.set()
 
@dp.message_handler(state=meinfo.Q1)                                
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    await state.update_data(answer1=answer)                            
    text = bold("Введите дату рождения") + "\nПример: " + "01.01.2000"
    await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
    await meinfo.Q2.set()                                   


@dp.message_handler(state=meinfo.Q2)                                
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    await state.update_data(answer2=answer)                            
    text = bold("Введите номер телефона")
    await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
    await meinfo.Q3.set()


@dp.message_handler(state=meinfo.Q3)                                
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    await state.update_data(answer3=answer)                            
    text = bold("Отправьте ссылку на VK")
    await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
    await meinfo.Q4.set() 

 

@dp.message_handler(state=meinfo.Q4)                                
async def answer_q1(message: types.Message, state: FSMContext):
    answer = message.text
    await state.update_data(answer4=answer)
    data = await state.get_data()               
    answer1 = data.get("answer1")
    answer2 = data.get("answer2") 
    answer3 = data.get("answer3")
    answer4 = data.get("answer4") 
    f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
    wb = load_workbook(f)
    ws = wb['Регистрация']
    ww = wb['Количество']  
    k = 0
    for i in range(2,10000):
            if str(ws['C' + str(i)].value) == str(answer1) and str(ws['D' + str(i)].value) == str(answer2) and str(ws['E' + str(i)].value) == str(answer3) and str(ws['F' + str(i)].value) == str(answer4):
                if str(ws['G' + str(i)].value) == '1':
                    text = "Вы в списках! Оплата успешно прошла!"
                    await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
                    k = 1
                    break
                    
                elif str(ws['G' + str(i)].value) == '0':
                    text = "Вы в списках!\n\nДля того, чтобы попасть на " + bold("HALLOWEEN PARTY") + ", необходимо оплатить билет.\nЕсли вы оплатили и прикрепили скрин в\n/payment, то ожидайте ответа в этом чате.\n" + bold('Вам придет сообщение')
                    await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
                    k = 1
                    break
    if k == 0:
        n =str(int(ww['A2'].value) + 2)
        ws['A' + n] = message.chat.id
        ws['B' + n] = message.from_user.username
        ws['C' + n] = answer1
        ws['D' + n] = answer2
        ws['E' + n] = answer3
        ws['F' + n] = answer4
        ws['G' + n] = 0
        ww['A2'].value = int(n)-1
        wb.save(f)
        wb.close()
        text = "Спасибо! Вы в списках!\n\nЧтобы попасть на " + bold("BEZNAZVANIYA") + ' | ' + bold('ACT A FOOL DAY') + ", необходимо оплатить билет.\n\nДля этого нажмите команду /payment или кнопку " + bold('"Оплата"') + " внизу👇"
        await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)
        J = open("/Users/senyashago/Desktop/Бот практика/bez/check.txt","r", encoding="utf-8") 
        joinedUsers = set()
        for line in J:
            joinedUsers.add(line.strip())
        J = open("/Users/senyashago/Desktop/Бот практика/bez/check.txt","a", encoding="utf-8")
        J.write(str(message.chat.id)+'\n')
        joinedUsers.add(message.chat.id)
    await state.finish()
    


@dp.message_handler(content_types=['text'], state=None)
async def enter_meinfo(message: types.Message):
    if message.text == 'Регистрация' or message.text == 'регистрация':  
        f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
        wb = load_workbook(f)
        ws = wb['Регистрация']
        text = bold("Введите ФИО") + "\nПример: " + italic("Иванов Иван Иванович")
        await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)
        await meinfo.Q1.set()



    elif message.text == 'Оплата' or message.text == "оплата":
        await message.answer(f'Выберите вариант оплаты', parse_mode=ParseMode.MARKDOWN, reply_markup=V_button)# \n/question - связаться с нами



    elif message.text == '1299 руб. с репостом афиши из группы VK к себе на страницу':
        J = open("/Users/senyashago/Desktop/Бот практика/bez/check.txt","r", encoding="utf-8")
        joinedUsers = set()
        for line in J:
            joinedUsers.add(line.strip())
        if str(message.chat.id) in joinedUsers:
            text = bold("Введите ФИО, указанное при регистрации") + "\nПример: " + italic("Иванов Иван Иванович")
            await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)                                
            await PGroup.W3.set()
        else:
            text = "Сначала зарегистрируйся. Для этого жми команду /enroll или кнопку " + bold('"Регистрация"') + " внизу👇"
            await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)



    elif message.text == '1499 руб. без репоста':
        J = open("/Users/senyashago/Desktop/Бот практика/bez/check.txt","r", encoding="utf-8")
        joinedUsers = set()
        for line in J:
            joinedUsers.add(line.strip())
        if str(message.chat.id) in joinedUsers:
            text = bold("Введите ФИО, указанное при регистрации") + "\nПример: " + italic("Иванов Иван Иванович")
            await message.answer(text, reply_markup=cancle_button, parse_mode=ParseMode.MARKDOWN)                                
            await PGroup.W1.set()
        else:
            text = "Сначала зарегистрируйся. Для этого жми команду /enroll или кнопку " + bold('"Регистрация"') + " внизу👇"
            await message.answer(text, reply_markup=start_button, parse_mode=ParseMode.MARKDOWN)









    elif message.text == 'Наши контакты' or message.text == "наши контакты":
        keyboard = InlineKeyboardMarkup()
        url_button1 = InlineKeyboardButton(text="Менеджер VK", url="https://vk.com/beznazvaniya_manager")
        url_button5 = InlineKeyboardButton(text="Наш инстаграм", url="https://instagram.com/beznazvaniya_msk?igshid=YmMyMTA2M2Y=")
        url_button6 = InlineKeyboardButton(text="Группа VK", url="https://vk.com/beznazvaniya_msk")
        url_button7 = InlineKeyboardButton(text="Telegram-канал", url="https://t.me/joinchat/AAAAAEuYaZ8EnD84P-MUXQ")
        keyboard.add(url_button1)
        keyboard.add(url_button5)
        keyboard.add(url_button6)
        keyboard.add(url_button7)
        await bot.send_message(message.from_user.id, bold('Контакты'),reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN)
    elif (message.text == "Таблица" or message.text == "таблица") and int(message.chat.id) == int(admin_chat_id):
        markup = types.ReplyKeyboardRemove()
        await bot.send_document(admin_chat_id, document=open('/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx', 'rb'), reply_markup=markup)
    elif int(message.chat.id) == int(admin_chat_id):
        chatId = message.text.split(': ')[0]
        H = message.text.split(': ')[1]
        if chatId == 'проблема' or chatId == 'Проблема':
            text = bold('УВЕДОМЛЕНИЕ') + '\n' + _join('Здравствуйте\n') + message.text.split(': ')[2] + '\n' + italic('С уважением beznazvaniya')
            await bot.send_message(H, text, parse_mode=ParseMode.MARKDOWN)
        else:
            text = bold('УВЕДОМЛЕНИЕ') + '\n\n' + _join('Салют! Оплата билета на ') + bold(message.text.split(': ')[1]) + ' прошла успешно!\n\n' + 'В ожидании тусовки подписывайся:'
            keyboard = InlineKeyboardMarkup()
            url_button2 = InlineKeyboardButton(text="Группа VK", url="https://vk.com/beznazvaniya_msk")
            url_button2_1 = InlineKeyboardButton(text="Telegram-чат", url="https://t.me/+AUHgO5gpk0lhMzMy")
            url_button3 = InlineKeyboardButton(text="Telegram-канал", url="https://t.me/joinchat/AAAAAEuYaZ8EnD84P-MUXQ")
            url_button4 = InlineKeyboardButton(text="Наш инстаграм", url="https://instagram.com/beznazvaniya_msk?igshid=YmMyMTA2M2Y=")
            keyboard.add(url_button2)
            keyboard.add(url_button2_1)
            keyboard.add(url_button3)
            keyboard.add(url_button4)
            pop = message.text.split(': ')[1]
            f = '/Users/senyashago/Desktop/Бот практика/bez/users3.xlsx'
            wb = load_workbook(f)
            ws = wb['Регистрация']
            N = 0
            ww = wb['Количество']
            for i in range(2,10000):
                if ws['G'+str(i)].value == 1:
                    N += 1
            ww['B2'].value = N
            ft = PatternFill(fill_type='solid', fgColor='008000')
            font = Font(color="ffffff")
            for i in range(2,100000):
                if str(ws['C' + str(i)].value) == str(pop):
                    ws['G' + str(i)] = 1
                    ws['A' + str(i)].fill = ft
                    ws['B' + str(i)].fill = ft
                    ws['C' + str(i)].fill = ft
                    ws['D' + str(i)].fill = ft
                    ws['E' + str(i)].fill = ft
                    ws['F' + str(i)].fill = ft
                    ws['G' + str(i)].fill = ft
                    ws['A' + str(i)].font = font
                    ws['B' + str(i)].font = font
                    ws['C' + str(i)].font = font
                    ws['D' + str(i)].font = font
                    ws['E' + str(i)].font = font
                    ws['F' + str(i)].font = font
                    ws['G' + str(i)].font = font
                    break
            wb.save(f)
            wb.close()
            await bot.send_message(chatId, text, reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN)



if __name__ == '__main__':
    executor.start_polling(dp)

